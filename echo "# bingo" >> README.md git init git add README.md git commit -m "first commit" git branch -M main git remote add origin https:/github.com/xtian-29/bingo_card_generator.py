import random
import xlsxwriter

class BingoCardGenerator:
    def __init__(self):
        self.card = [[None for _ in range(5)] for _ in range(5)]
        self.numbers = []

    def generate_card(self):
        self._generate_bingo_letters()
        self._generate_numbers()
        self._populate_card()

    def _generate_bingo_letters(self):
        letters = list('BINGO')
        random.shuffle(letters)
        for i in range(5):
            self.card[0][i] = letters[i]

    def _generate_numbers(self):
        self.numbers = random.sample(range(1, 76), 25)
        self._rearrange_numbers()

    def _rearrange_numbers(self):
        for i in range(5):
            column_numbers = [n for n in self.numbers[i * 5:(i * 5) + 5]]
            if all(n % 2 == 0 for n in column_numbers):
                random.shuffle(column_numbers)
                while all(n % 2 == 0 for n in column_numbers):
                    random.shuffle(column_numbers)
            elif all(n % 2 != 0 for n in column_numbers):
                random.shuffle(column_numbers)
                while all(n % 2 != 0 for n in column_numbers):
                    random.shuffle(column_numbers)
            self.numbers[i * 5:(i * 5) + 5] = column_numbers

    def _populate_card(self):
        used_indices = set()
        for i in range(5):
            used_indices.add(i * 5 + random.randint(0, 4))
            self.card[i][self.card[i].index(None)] = self.numbers[i * 5:(i * 5) + 5][0]
        for i in range(5):
            for j in range(5):
                if self.card[j][i] is None and (i * 5 + j) not in used_indices:
                    self.card[j][i] = self.numbers[i * 5:(i * 5) + 5][j]

    def export_cards_to_excel(self, num_batches):
        workbook = xlsxwriter.Workbook('bingo_cards.xlsx')
        for i in range(num_batches):
            worksheet = workbook.add_worksheet(f'Batch {i+1}')
            for j in range(500):
                self.generate_card()
                for k in range(5):
                    worksheet.write(j, k, self.card[0][k])
                for k in range(5):
                    worksheet.write(j, k+5, self.card[1][k])
                for k in range(5):
                    worksheet.write(j, k+10, self.card[2][k])
                for k in range(5):
                    worksheet.write(j, k+15, self.card[3][k])
                for k in range(5):
                    worksheet.write(j, k+20, self.card[4][k])
        workbook.close()
import sys
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QLabel, QMainWindow, QScrollBar, QTextEdit, QVBoxLayout, QWidget

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # set window title
        self.setWindowTitle("Bingo Card Generator")

        # set window size
        self.setFixedSize(800, 600)

        # create central widget and layout
        central_widget = QWidget()
        main_layout = QVBoxLayout()

        # create scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area_content = QWidget(scroll_area)
        scroll_area.setWidget(scroll_area_content)
        scroll_area_layout = QVBoxLayout(scroll_area_content)

        # create label for the bingo letters
        bingo_letters = QLabel()
        bingo_letters.setText("B   I   N   G   O")
        bingo_letters.setFont(QFont('Arial', 20))
        bingo_letters.setAlignment(Qt.AlignCenter)

        # add bingo letters to main layout
        main_layout.addWidget(bingo_letters)

        # add scroll area to main layout
        main_layout.addWidget(scroll_area)

        # create 500 bingo cards
        cards = generate_cards()

        # add bingo cards to scroll area layout
        for i, card in enumerate(cards):
            card_widget = CardWidget(card)
            scroll_area_layout.addWidget(card_widget)

            # only show the first 25 cards
            if i >= 24:
                card_widget.hide()

        # create a label to show the called number
        called_number_label = QLabel()
        called_number_label.setFont(QFont('Arial', 14))
        called_number_label.setAlignment(Qt.AlignCenter)

        # add called number label to main layout
        main_layout.addWidget(called_number_label)

        # set central widget layout
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # connect card widgets to signal for checking if a number is on the card
        for card_widget in scroll_area_content.findChildren(CardWidget):
            card_widget.number_checked.connect(self.number_checked)

        # start the game
        self.cards = cards
        self.current_card_index = 0
        self.called_numbers = []
        self.called_number_label = called_number_label
        self.next_turn()

    def number_checked(self, number, found):
        if found:
            self.called_number_label.setText(f"Number called: {number} - FOUND on a card!")
        else:
            self.called_number_label.setText(f"Number called: {number}")

    def next_turn(self):
        # get a new number and add it to the called numbers list
        new_number = generate_number(self.called_numbers)
        self.called_numbers.append(new_number)
        self.called_number_label.setText(f"Number called: {new_number}")

        # check if the number is on any of the cards
        for i in range(self.current_card_index, len(self.cards)):
            card_widget = self.findChild(CardWidget, f"card{i}")
            found = card_widget.check_number(new_number)
            if found:
                self.current_card_index = i
                card_widget.setStyleSheet("background-color: yellow")
                card_widget.show()
                break

        # check if game is over
        if self.current_card_index == len(self.cards) - 1:
            reply = QMessageBox.question(self, "Game Over", "All cards have been played. Do you want to play again?",
                                          QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.restart_game()
            else:
                self.close()
        else:
            # show the next card
            next_card_widget = self.findChild(CardWidget, f"card{self.current_card_index + 1}")
            next_card_widget.show()

    def restart_game(self):
        # create 500 new cards
        new_cards = generate_cards
class BingoWindow(QMainWindow):
    def __init__(self, parent=None):
        super(BingoWindow, self).__init__(parent)
        self.setWindowTitle("Bingo Game")
        self.setGeometry(100, 100, 800, 600)
        self.setMinimumSize(600, 400)
        self.setStyleSheet("background-color: white;")
        self.bingo_batch = []
        self.card_batch = []
        self.current_card_batch = []
        self.current_index = 0
        self.max_displayed_cards = 25
        self.number_called = ''
        self.called_numbers = []

        # Menu Bar
        menu_bar = self.menuBar()
        menu_file = menu_bar.addMenu("File")
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        menu_file.addAction(exit_action)

        # Status Bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # Central Widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QHBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        # Bingo Numbers Display
        self.numbers_display_layout = QVBoxLayout()
        self.numbers_display_layout.setSpacing(0)
        self.numbers_display_layout.setAlignment(Qt.AlignTop)
        self.main_layout.addLayout(self.numbers_display_layout)

        self.current_number_label = QLabel("Current Number: ")
        self.current_number_label.setAlignment(Qt.AlignCenter)
        self.numbers_display_layout.addWidget(self.current_number_label)

        self.called_numbers_label = QLabel("Called Numbers:")
        self.called_numbers_label.setAlignment(Qt.AlignCenter)
        self.numbers_display_layout.addWidget(self.called_numbers_label)

        self.called_numbers_list = QListWidget()
        self.called_numbers_list.setFixedSize(150, 400)
        self.numbers_display_layout.addWidget(self.called_numbers_list)

        # Bingo Cards Display
        self.cards_display_layout = QVBoxLayout()
        self.cards_display_layout.setAlignment(Qt.AlignTop)
        self.main_layout.addLayout(self.cards_display_layout)

        self.cards_batch_label = QLabel("Cards Batch")
        self.cards_batch_label.setAlignment(Qt.AlignCenter)
        self.cards_display_layout.addWidget(self.cards_batch_label)

        self.cards_scroll_area = QScrollArea()
        self.cards_scroll_area.setWidgetResizable(True)
        self.cards_display_layout.addWidget(self.cards_scroll_area)

        self.cards_widget = QWidget()
        self.cards_widget_layout = QGridLayout()
        self.cards_widget.setLayout(self.cards_widget_layout)
        self.cards_scroll_area.setWidget(self.cards_widget)

        # Buttons
        self.buttons_layout = QVBoxLayout()
        self.buttons_layout.setAlignment(Qt.AlignTop)
        self.main_layout.addLayout(self.buttons_layout)

        self.next_button = QPushButton("Next")
        self.next_button.setFixedSize(150, 50)
        self.next_button.clicked.connect(self.next_card)
        self.buttons_layout.addWidget(self.next_button)

        self.reset_button = QPushButton("Reset")
        self.reset_button.setFixedSize(150, 50)
        self.reset_button.clicked.connect(self.reset_game)
        self.buttons_layout.addWidget(self.reset_button)

        # Bingo Card Generator
        self.generate_card_batch()

        # Update UI
        self.update_current_number_label()

    def generate_card_batch(self):
        self.card_batch = generate_card_numbers(500)

        # Reset Current Card Batch
        self.current_card_batch = self.card_batch[:self.max_displayed_cards]
        self.current_index = 0

        # Clear Current Displayed Cards
        for i in reversed(range(self.cards_widget_layout.count())):
            self.cards_widget_layout.itemAt(i).widget().setParent(None)

        # Display Current Card Batch
        for i in range(len(self.current_card_batch)):
            card_layout = QVBoxLayout()
import random
import xlsxwriter

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QLabel, QMainWindow, QScrollBar, QSizePolicy


class BingoCard:
    def __init__(self):
        self.card = []
        self.generate_card_numbers()

    def generate_card_numbers(self):
        numbers = {"B": [], "I": [], "N": [], "G": [], "O": []}
        for key in numbers.keys():
            for i in range(5):
                if key == "N" and i == 2:
                    continue
                while True:
                    num = random.randint(i * 15 + 1, i * 15 + 15)
                    if num not in numbers[key]:
                        if (i == 0 and num <= 15) or (i == 4 and num >= 61):
                            continue
                        if i > 0 and num in numbers[key][i - 1]:
                            continue
                        numbers[key].append(num)
                        break
        self.card = [[0] * 5 for i in range(5)]
        for key, vals in numbers.items():
            for i, val in enumerate(vals):
                self.card[i][{"B": 0, "I": 1, "N": 2, "G": 3, "O": 4}[key]] = val


class BingoBatch:
    def __init__(self, num_cards):
        self.cards = []
        self.generate_cards(num_cards)

    def generate_cards(self, num_cards):
        for i in range(num_cards):
            card = BingoCard()
            self.cards.append(card)

    def get_card_numbers(self):
        card_numbers = []
        for card in self.cards:
            card_numbers.append([card.card[i][j] for j in range(5) for i in range(5)])
        return card_numbers


class BingoCaller:
    def __init__(self):
        self.called_numbers = []

    def call_number(self):
        while True:
            num = random.randint(1, 75)
            if num not in self.called_numbers:
                self.called_numbers.append(num)
                return num


class BingoWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.batch = None
        self.num_called = 0
        self.bingo_caller = BingoCaller()
        self.setWindowTitle("Bingo")
        self.setGeometry(100, 100, 600, 800)
        self.central_widget = QLabel(self)
        self.central_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.central_widget.setAlignment(Qt.AlignTop)
        self.setCentralWidget(self.central_widget)
        self.scrollbar = QScrollBar(Qt.Vertical)
        self.scrollbar.setPageStep(25)
        self.scrollbar.valueChanged.connect(self.update_widget)
        self.scrollbar.setMaximum(0)
        self.scrollbar.setMinimum(0)
        self.scrollbar.setValue(0)
        self.scrollbar.setTracking(True)
        self.numbers_label = QLabel(self.central_widget)
        self.numbers_label.setGeometry(0, 0, 600, 50)
        font = QFont()
        font.setPointSize(16)
        self.numbers_label.setFont(font)
        self.numbers_label.setAlignment(Qt.AlignCenter)
        self.update_numbers_label()
        self.show()

    def update_widget(self):
        value = self.scrollbar.value()
        self.central_widget.setGeometry(0, -value, 600, 800)

    def update_numbers_label(self):
        self.numbers_label.setText(f"Number called: {self.bingo_caller
import random
import xlsxwriter

BINGO_COLUMNS = ['B', 'I', 'N', 'G', 'O']
NUM_CARDS = 500
NUM_BATCHES = 1
CARD_SIZE = 5
NUM_PER_COL = 15
NUM_ROWS = 5
NUM_COLS = 5
NUMBERS_PER_LETTER = 15

def generate_card_numbers():
    card = []
    for letter in BINGO_COLUMNS:
        used_numbers = []
        for i in range(NUM_PER_COL):
            valid_numbers = set(range((BINGO_COLUMNS.index(letter) * NUMBERS_PER_LETTER) + 1,
                                       (BINGO_COLUMNS.index(letter) * NUMBERS_PER_LETTER) + NUMBERS_PER_LETTER + 1))
            valid_numbers -= set(used_numbers)
            valid_numbers -= set(used_numbers[i-1:i+1])
            number = random.choice(list(valid_numbers))
            used_numbers.append(number)
            card.append(number)
    return card

def generate_cards(num_cards):
    cards = []
    for i in range(num_cards):
        card = generate_card_numbers()
        cards.append(card)
    return cards

def export_cards_to_excel(cards):
    workbook = xlsxwriter.Workbook('Bingo Cards.xlsx')
    for i, card_batch in enumerate(cards):
        worksheet = workbook.add_worksheet(f'Batch {i+1}')
        for row in range(NUM_ROWS):
            for col in range(NUM_COLS):
                cell_value = card_batch[NUM_COLS * row + col]
                worksheet.write(row, col, cell_value)
    workbook.close()

cards = generate_cards(NUM_CARDS * NUM_BATCHES)
export_cards_to_excel(cards)
print(f'Exported {NUM_CARDS * NUM_BATCHES} cards to Bingo Cards.xlsx')
class BingoWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.batch_number = 1
        self.cards_per_batch = 500
        self.cards = []

        self.setWindowTitle("Bingo")
        self.setGeometry(50, 50, 1100, 700)
        self.create_bingo_card()

        self.caller_label = QLabel("", self)
        self.caller_label.setGeometry(600, 30, 150, 50)
        font = QFont()
        font.setPointSize(20)
        font.setBold(True)
        self.caller_label.setFont(font)
        self.caller_label.setAlignment(Qt.AlignCenter)

        self.batch_label = QLabel("", self)
        self.batch_label.setGeometry(800, 30, 150, 50)
        font = QFont()
        font.setPointSize(20)
        font.setBold(True)
        self.batch_label.setFont(font)
        self.batch_label.setAlignment(Qt.AlignCenter)

        self.timer_label = QLabel("", self)
        self.timer_label.setGeometry(1000, 30, 150, 50)
        font = QFont()
        font.setPointSize(20)
        font.setBold(True)
        self.timer_label.setFont(font)
        self.timer_label.setAlignment(Qt.AlignCenter)

        self.call_button = QPushButton("Call", self)
        self.call_button.setGeometry(600, 100, 100, 50)
        self.call_button.clicked.connect(self.call_number)

        self.generate_cards_button = QPushButton("Generate Cards", self)
        self.generate_cards_button.setGeometry(800, 100, 150, 50)
        self.generate_cards_button.clicked.connect(self.generate_cards)

        self.export_button = QPushButton("Export Cards", self)
        self.export_button.setGeometry(1000, 100, 150, 50)
        self.export_button.clicked.connect(self.export_cards)

        self.scroll_area = QScrollArea(self)
        self.scroll_area.setGeometry(50, 200, 1000, 400)

        self.scroll_widget = QWidget()
        self.scroll_widget.setGeometry(0, 0, 1000, 600)

        self.scroll_layout = QGridLayout(self.scroll_widget)
        self.scroll_layout.setSpacing(20)

        self.scroll_area.setWidget(self.scroll_widget)
        self.scroll_area.setWidgetResizable(True)

    def create_bingo_card(self):
        self.cards = generate_card_numbers(self.cards_per_batch)
        for i in range(self.cards_per_batch):
            card_widget = QWidget()
            card_widget.setFixedSize(250, 250)
            card_layout = QGridLayout(card_widget)
            card_layout.setSpacing(10)

            # Set the Bingo letters
            b_label = QLabel("B", card_widget)
            i_label = QLabel("I", card_widget)
            n_label = QLabel("N", card_widget)
            g_label = QLabel("G", card_widget)
            o_label = QLabel("O", card_widget)

            b_label.setAlignment(Qt.AlignCenter)
            i_label.setAlignment(Qt.AlignCenter)
            n_label.setAlignment(Qt.AlignCenter)
            g_label.setAlignment(Qt.AlignCenter)
            o_label.setAlignment(Qt.AlignCenter)

            font = QFont()
            font.setPointSize(16)
            font.setBold(True)
            b_label.setFont(font)
            i_label.setFont(font)
            n_label.setFont(font)
            g_label.setFont(font)
            o_label.setFont(font)

            card_layout.addWidget(b_label, 0, 0)
            card_layout.addWidget(i_label, 0, 1)
            card_layout.addWidget(n_label, 0, 2)
            card_layout.addWidget(g_label, 0, 3)
            card_layout.addWidget(o_label, 0, 4
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QWidget, QGridLayout, QScrollArea
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from random import sample, randint
import pandas as pd


class GenerateCardNumbers(QThread):
    signal_cards_generated = pyqtSignal(list)

    def __init__(self):
        super().__init__()

    def run(self):
        numbers = {'B': [], 'I': [], 'N': [], 'G': [], 'O': []}
        for letter, col in numbers.items():
            i = 1
            while len(col) < 5:
                if letter == 'N' and i == 3:
                    col.append('')
                else:
                    col.append(i + (15 * ('BINGO'.index(letter))))
                    i += 1

        def get_random_numbers():
            available_numbers = []
            for i in range(1, 76):
                if i not in used_numbers:
                    available_numbers.append(i)
            return sample(available_numbers, 15)

        def check_valid(card, num):
            if card[num] != '':
                return False
            col = num // 15
            if num % 15 == 0:
                count = len([i for i in card[col * 15:col * 15 + 5] if i != ''])
                if count == 5:
                    return False
            elif num % 15 == 14:
                count = len([i for i in card[col * 15 + 10:col * 15 + 15] if i != ''])
                if count == 5:
                    return False
            else:
                count_left = len([i for i in card[col * 15:num] if i != ''])
                count_right = len([i for i in card[num + 1:col * 15 + 15] if i != ''])
                if count_left == 4 and count_right == 4:
                    return False
            row_start = col * 15
            if num < row_start + 5:
                count = len([i for i in card[num % 15::15] if i != ''])
                if count == 5:
                    return False
            if num >= row_start + 10:
                count = len([i for i in card[(num % 15) + 10::15] if i != ''])
                if count == 5:
                    return False
            if num == row_start + 2:
                if '' in card[row_start:row_start + 5]:
                    return False
            used_in_col = [i for i in range(col * 15, col * 15 + 15) if card[i] != '']
            if len(used_in_col) == 5 and used_in_col[0] + 15 == used_in_col[-1]:
                return False
            if all(i % 2 == 0 for i in used_in_col) or all(i % 2 == 1 for i in used_in_col):
                return False
            return True

        card_numbers = []
        for _ in range(500):
            card = [''] * 25
            used_numbers = []
            for i, letter in enumerate(numbers):
                available_spaces = sample([x for x in range(i * 5, (i + 1) * 5)], 3)
                for num in available_spaces:
                    while True:
                        num_to_add = randint(i * 15 + 1,
class BingoCard:
    def __init__(self, card_id, letters):
        self.card_id = card_id
        self.letters = letters
        self.card = []
        self.generate_card_numbers()

    def generate_card_numbers(self):
        for letter in self.letters:
            self.card.extend(self.get_numbers_for_letter(letter))

    def get_numbers_for_letter(self, letter):
        numbers = []
        if letter == "B":
            numbers = random.sample(range(1, 16), 5)
        elif letter == "I":
            numbers = random.sample(range(16, 31), 5)
        elif letter == "N":
            numbers = random.sample(range(31, 46), 4)
            numbers.insert(2, 0)
        elif letter == "G":
            numbers = random.sample(range(46, 61), 5)
        elif letter == "O":
            numbers = random.sample(range(61, 76), 5)

        return numbers

    def __str__(self):
        card_str = f"Card ID: {self.card_id}\n"
        for i in range(5):
            row = self.card[i*5:(i+1)*5]
            card_str += f"{self.letters[i]} | {row[0]:2} | {row[1]:2} | {row[2]:2} | {row[3]:2} | {row[4]:2}\n"
        return card_str


class BingoBatch:
    def __init__(self, batch_id, batch_size=500):
        self.batch_id = batch_id
        self.batch_size = batch_size
        self.cards = []
        self.generate_batch()

    def generate_batch(self):
        for i in range(self.batch_size):
            card_id = f"{self.batch_id}-{i+1}"
            letters = ["B", "I", "N", "G", "O"]
            random.shuffle(letters)
            while not self.validate_card(letters):
                random.shuffle(letters)
            card = BingoCard(card_id, letters)
            self.cards.append(card)

    def validate_card(self, letters):
        for i in range(5):
            column = []
            for j in range(5):
                if letters[j] == self.get_letter_for_column(i):
                    column.append(j)
            if len(column) == 5:
                return False
            if len(column) > 1:
                for k in range(len(column)-1):
                    if abs(column[k] - column[k+1]) == 1:
                        return False
            if len(column) == 4:
                if column[0] == 0 or column[-1] == 4:
                    return False
                if all(self.is_odd(letters[j]) for j in column) or all(self.is_even(letters[j]) for j in column):
                    return False
        return True

    def get_letter_for_column(self, column):
        return ["B", "I", "N", "G", "O"][column]

    def is_odd(self, number):
        return number % 2 == 1

    def is_even(self, number):
        return number % 2 == 0


class BingoGame:
    def __init__(self):
        self.batch_size = 500
        self.batches = []
        self.generate_batches()
        self.current_batch_index = 0
        self.called_numbers = set()
        self.call_next_number()

    def generate_batches(self):
        for i in range
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QHBoxLayout, QWidget, QScrollArea
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont
from random import shuffle, randrange
from itertools import product
from openpyxl import Workbook


class BingoCard:
    def __init__(self):
        self.card = None
        self.generate_card()

    def generate_card(self):
        valid_card = False
        while not valid_card:
            self.card = []
            for i in range(5):
                column = []
                if i == 2:
                    column.append("FREE")
                    for j in range(4):
                        rand_num = randrange(j * 15 + 1, (j + 1) * 15 + 1)
                        column.append(rand_num)
                else:
                    for j in range(5):
                        rand_num = randrange(j * 15 + 1, (j + 1) * 15 + 1)
                        column.append(rand_num)
                shuffle(column)
                self.card.append(column)
            valid_card = self.check_card_validity()

    def check_card_validity(self):
        # Check for consecutive numbers in the same column
        for col in self.card:
            for i in range(len(col) - 1):
                if col[i+1] - col[i] == 1:
                    return False

        # Check for all odds or all evens in the same column
        for col in self.card:
            if all(num % 2 == 0 for num in col) or all(num % 2 == 1 for num in col):
                return False

        return True


class BingoBatch:
    def __init__(self, batch_size=500):
        self.batch = []
        self.batch_size = batch_size
        self.generate_batch()

    def generate_batch(self):
        while len(self.batch) < self.batch_size:
            card = BingoCard()
            if card.card not in self.batch:
                self.batch.append(card.card)

    def export_to_excel(self, filename="bingo_cards.xlsx"):
        wb = Workbook()
        ws = wb.active
        for i, card in enumerate(self.batch):
            col_offset = i % 10 * 7
            row_offset = (i // 10) * 6
            for row, col in product(range(6), range(5)):
                cell_value = card[col][row] if col != 2 or row != 2 else ""
                cell = ws.cell(row=row + 1 + row_offset, column=col + 1 + col_offset, value=cell_value)
                if col == 2 and row == 2:
                    cell.style = "CenteredText"
            ws.cell(row=7 + row_offset, column=6 + col_offset, value="B")
            ws.cell(row=7 + row_offset, column=7 + col_offset, value="I")
            ws.cell(row=7 + row_offset, column=8 + col_offset, value="N")
            ws.cell(row=7 + row_offset, column=9 + col_offset, value="G")
            ws.cell(row=7 + row_offset, column=10 + col_offset, value="O")
        wb.save(filename)


class BingoWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Bingo Card Generator")
        self.setFixedSize(800, 600)

        # Generate batch of bingo cards and export to Excel
        self.bingo_batch = BingoBatch(batch_size=500)
       
import random
import xlsxwriter
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QScrollArea, QWidget, QVBoxLayout
from PyQt5.QtCore import Qt, pyqtSignal


class BingoCardGenerator:
    def __init__(self):
        self.cards = []
        self.batch_size = 500

    def generate_card_numbers(self):
        self.cards = []
        while len(self.cards) < self.batch_size:
            card = []
            b_col = random.sample(range(1, 16), 5)
            i_col = random.sample(range(16, 31), 5)
            n_col = random.sample(range(31, 46), 5)
            g_col = random.sample(range(46, 61), 5)
            o_col = random.sample(range(61, 76), 5)
            for i in range(5):
                card.append(b_col[i])
                card.append(i_col[i])
                card.append(n_col[i] if i != 2 else ' ')
                card.append(g_col[i])
                card.append(o_col[i])
            if self.check_card_validity(card):
                self.cards.append(card)

    def check_card_validity(self, card):
        for i in range(5):
            if card[i] % 2 == 0 and card[i + 5] % 2 == 0 and card[i + 10] % 2 == 0 and card[i + 15] % 2 == 0 and card[
                i + 20] % 2 == 0:
                return False
            if card[i] % 2 == 1 and card[i + 5] % 2 == 1 and card[i + 10] % 2 == 1 and card[i + 15] % 2 == 1 and card[
                i + 20] % 2 == 1:
                return False
            if i < 4:
                if card[i * 5] + 1 == card[i * 5 + 1] == card[i * 5 + 2] == card[i * 5 + 3] + 1 or card[
                    i * 5] - 1 == card[i * 5 + 1] == card[i * 5 + 2] == card[i * 5 + 3] - 1:
                    return False
        return True

    def export_cards_to_excel(self):
        workbook = xlsxwriter.Workbook('bingo_cards.xlsx')
        worksheet = workbook.add_worksheet()

        row = 0
        col = 0

        for card in self.cards:
            for number in card:
                worksheet.write(row, col, number)
                col += 1
                if col == 25:
                    row += 1
                    col = 0
            row += 1
            col = 0

        workbook.close()


class BingoWindow(QMainWindow):
    my_signal = pyqtSignal(list)

    def __init__(self):
        super().__init__()

        self.setWindowTitle("Bingo Card Generator")

        # Set up card generator
        self.generator = BingoCardGenerator()
        self.generator.generate_card_numbers()
        self.generator.export_cards_to_excel()

        # Set up main widget
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)

        # Set up vertical layout
        self.layout = QVBoxLayout(self.main_widget)

        # Set up scroll area
        self.scroll_area = QScrollArea(self)
        self.scroll

import openpyxl

# create a new workbook
workbook = openpyxl.Workbook()

# select the active worksheet
worksheet = workbook.active

# write data to the worksheet
worksheet['A1'] = 'Bingo Card 1'
worksheet['A2'] = 'B'
worksheet['B2'] = 'I'
worksheet['C2'] = 'N'
worksheet['D2'] = 'G'
worksheet['E2'] = 'O'

# save the workbook
workbook.save('bingo_cards.xlsx')

